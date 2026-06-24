#!/usr/bin/env python3
"""
to_excel.py — Stage 4. Build a formatted workbook from rows.json.

Usage:
    python3 to_excel.py --work WORKDIR --out Madison_EMS_Schedules.xlsx \
        --pdf-name Madison_EMS_Schedules_Jun2025_Jun2026.pdf [--source-note "..."]
"""
import argparse, json, os, re
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLS = ['Date','Day','Unit','Employee','Start Time','End Time','Entry Type','Notes','PDF Reference']
VEH = re.compile(r'^(medic|ambulance)\s*\d{2,3}$', re.I)
ROLES = {'paramedic','emt','paramedic/emt','emt/paramedic'}


def clean(df):
    df['Employee'] = df['Employee'].astype(str).str.strip()
    df['Notes'] = df['Notes'].astype(str).str.replace('nan', '', regex=False).str.strip()
    df = df[~df['Employee'].str.match(VEH)].copy()
    def relabel(r):
        emp = r['Employee']; low = emp.lower()
        if low in ROLES or emp in ('Surf Club',):
            tag = f'Unfilled slot ({emp})' if low in ROLES else f'Event: {emp}'
            r['Notes'] = tag + (' ; ' + r['Notes'] if r['Notes'] else '')
            r['Employee'] = '(unfilled)'
        return r
    return df.apply(relabel, axis=1).reset_index(drop=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--work', required=True)
    ap.add_argument('--out', default='Madison_EMS_Schedules.xlsx')
    ap.add_argument('--pdf-name', default='source.pdf')
    ap.add_argument('--source-note', default='', help='extra note shown on the Source sheet')
    ap.add_argument('--orig-page-offset', type=int, default=0,
                    help='if this PDF was extracted from a larger doc, add N to get the original page #')
    a = ap.parse_args()

    rows = json.load(open(os.path.join(a.work, 'rows.json')))
    keep = COLS[:-1] + ['_p']
    df = pd.DataFrame(rows)[keep].copy()
    df['Date'] = pd.to_datetime(df['Date'])
    df = clean(df)
    df = df.sort_values(['Date','Unit','Employee'], kind='stable').reset_index(drop=True)

    off = a.orig_page_offset
    def ref(r):
        if 'OCR uncertain' not in (r['Notes'] or ''):
            return ''
        p = int(r['_p']); d = pd.to_datetime(r['Date'])
        unit = (r['Unit'] or '').strip() or r['Entry Type']
        orig = f' (source doc p.{p + off})' if off else ''
        return f"{a.pdf_name}, p.{p}{orig} — {d.strftime('%a %m/%d/%Y')}, {unit} cell"
    df['PDF Reference'] = df.apply(ref, axis=1)
    df = df[COLS]
    ncol = len(COLS); nrows = len(df)

    with pd.ExcelWriter(a.out, engine='openpyxl', datetime_format='mm/dd/yyyy') as xw:
        df.to_excel(xw, index=False, sheet_name='Schedule')
        ws = xw.sheets['Schedule']
        hfill = PatternFill('solid', fgColor='1F4E78'); hfont = Font(bold=True, color='FFFFFF', size=11)
        thin = Side(style='thin', color='D0D0D0'); border = Border(thin, thin, thin, thin)
        widths = {'Date':12,'Day':6,'Unit':14,'Employee':22,'Start Time':10,'End Time':10,
                  'Entry Type':14,'Notes':40,'PDF Reference':56}
        for ci, name in enumerate(COLS, 1):
            c = ws.cell(1, ci); c.fill = hfill; c.font = hfont
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(ci)].width = widths[name]
        type_fill = {'Trade':'FFF2CC','Time Off':'FCE4D6','Vacation':'E2EFDA','Personal Time':'DDEBF7',
                     'Booked Shift':'EDEDED','Events Crew':'F4CCCC'}
        ref_fill = PatternFill('solid', fgColor='FFF2CC')
        for r in range(2, nrows + 2):
            fc = type_fill.get(ws.cell(r, 7).value)
            for ci in range(1, ncol + 1):
                c = ws.cell(r, ci); c.border = border
                c.alignment = Alignment(vertical='top', wrap_text=(ci in (8, 9)),
                                        horizontal='center' if ci in (2,5,6,7) else 'left')
                if fc: c.fill = PatternFill('solid', fgColor=fc)
            if ws.cell(r, 9).value: ws.cell(r, 9).fill = ref_fill
            ws.cell(r, 1).number_format = 'mm/dd/yyyy'
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:{get_column_letter(ncol)}{nrows+1}'

        unc = int(df['PDF Reference'].astype(str).str.len().gt(0).sum())
        info = [
            ('Madison EMS Schedules — extracted dataset', ''),
            ('', ''),
            ('Source file', a.pdf_name + ((' — ' + a.source_note) if a.source_note else '')),
            ('Coverage', f"{df['Date'].min().date():%m/%d/%Y} – {df['Date'].max().date():%m/%d/%Y}"),
            ('Total rows', str(nrows)),
            ('Distinct dates', str(df['Date'].dt.date.nunique())),
            ('Distinct people', str(df[df['Employee'] != '(unfilled)']['Employee'].nunique())),
            ('', ''),
            ('One row per', 'person per shift entry'),
            ('Entry Type values', 'Regular, Trade, Time Off, Booked Shift, Events Crew, Vacation, Personal Time'),
            ('Overnight shifts', '19:00-07:00 span midnight; 24h shifts show as 07:00-07:00 (one row)'),
            ('PDF Reference', f'populated for the {unc} cells flagged "OCR uncertain" — page + date/cell to verify by hand'),
            ('', ''),
            ('Caveats', ''),
            (f'  • {unc} rows "OCR uncertain"', 'source text partly obscured (edge clips / tooltip overlays); best-read values.'),
            ('  • "marked *" in Notes', 'a small asterisk/arrow by the time (often a partial/swap shift).'),
            ('  • "(unfilled)" Employee', 'an open Events-Crew slot / role placeholder, not a named person.'),
        ]
        pd.DataFrame(info, columns=['Field','Detail']).to_excel(xw, index=False, sheet_name='Source & Notes')
        sws = xw.sheets['Source & Notes']
        sws.column_dimensions['A'].width = 26; sws.column_dimensions['B'].width = 100
        sws.cell(1, 1).font = Font(bold=True, size=13); sws.cell(1, 2).value = None
        for r in range(2, len(info) + 2):
            if sws.cell(r, 1).value: sws.cell(r, 1).font = Font(bold=True)
            sws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical='top')
        xw.book.move_sheet('Source & Notes', -(len(xw.book._sheets) - 1))
    print(f'wrote {a.out}: {nrows} rows, {unc} PDF references')


if __name__ == '__main__':
    main()
